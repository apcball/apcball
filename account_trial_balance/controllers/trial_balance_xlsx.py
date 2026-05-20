# -*- coding: utf-8 -*-
import io

from odoo import http
from odoo.http import request

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class TrialBalanceXlsxController(http.Controller):

    @http.route(
        '/account_trial_balance/xlsx/<int:wizard_id>',
        type='http', auth='user',
    )
    def download_trial_balance_xlsx(self, wizard_id):
        wizard = request.env['account.trial.balance.wizard'].browse(wizard_id)
        wizard.ensure_one()

        data = wizard._get_trial_balance_data()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Trial Balance'

        # --- Styles ---
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')

        group_font = Font(name='Calibri', size=11, bold=True, color='2C3E50')
        group_fill = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')

        total_font = Font(name='Calibri', size=11, bold=True)
        total_fill = PatternFill(start_color='EAF2F8', end_color='EAF2F8', fill_type='solid')

        grand_font = Font(name='Calibri', size=12, bold=True, color='2C3E50')
        grand_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')

        data_font = Font(name='Calibri', size=10)
        data_align = Alignment(vertical='center')
        money_align = Alignment(horizontal='right', vertical='center')

        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 20

        # --- Title ---
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = 'Trial Balance — %s to %s' % (wizard.date_from, wizard.date_to)
        title_cell.font = Font(name='Calibri', size=14, bold=True, color='2C3E50')
        title_cell.alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:E2')
        subtitle = ws['A2']
        subtitle.value = 'Journals: %s' % (
            ', '.join(wizard.journal_ids.mapped('code')) if wizard.journal_ids else 'All'
        )
        subtitle.font = Font(name='Calibri', size=10, color='777777')
        subtitle.alignment = Alignment(horizontal='center')

        # --- Headers (row 4) ---
        headers = ['Code', 'Account', 'Debit', 'Credit', 'Balance']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        row = 5

        for group in data['lines']:
            # Group header row
            ws.cell(row=row, column=1, value=group['type']).font = group_font
            ws.cell(row=row, column=1).fill = group_fill
            ws.cell(row=row, column=1).border = thin_border
            for c in range(2, 6):
                cell = ws.cell(row=row, column=c)
                cell.fill = group_fill
                cell.border = thin_border
            row += 1

            # Account lines
            for line in group['lines']:
                ws.cell(row=row, column=1, value=line['code']).font = data_font
                ws.cell(row=row, column=1).alignment = data_align
                ws.cell(row=row, column=1).border = thin_border

                ws.cell(row=row, column=2, value=line['name']).font = data_font
                ws.cell(row=row, column=2).alignment = data_align
                ws.cell(row=row, column=2).border = thin_border

                ws.cell(row=row, column=3, value=round(line['debit'], 2)).font = data_font
                ws.cell(row=row, column=3).alignment = money_align
                ws.cell(row=row, column=3).number_format = '#,##0.00'
                ws.cell(row=row, column=3).border = thin_border

                ws.cell(row=row, column=4, value=round(line['credit'], 2)).font = data_font
                ws.cell(row=row, column=4).alignment = money_align
                ws.cell(row=row, column=4).number_format = '#,##0.00'
                ws.cell(row=row, column=4).border = thin_border

                ws.cell(row=row, column=5, value=round(line['balance'], 2)).font = data_font
                ws.cell(row=row, column=5).alignment = money_align
                ws.cell(row=row, column=5).number_format = '#,##0.00'
                ws.cell(row=row, column=5).border = thin_border

                row += 1

            # Group total row
            ws.cell(row=row, column=1, value='').font = total_font
            ws.cell(row=row, column=1).fill = total_fill
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value='Total %s' % group['type']).font = total_font
            ws.cell(row=row, column=2).fill = total_fill
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=round(group['total_debit'], 2)).font = total_font
            ws.cell(row=row, column=3).fill = total_fill
            ws.cell(row=row, column=3).number_format = '#,##0.00'
            ws.cell(row=row, column=3).alignment = money_align
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4, value=round(group['total_credit'], 2)).font = total_font
            ws.cell(row=row, column=4).fill = total_fill
            ws.cell(row=row, column=4).number_format = '#,##0.00'
            ws.cell(row=row, column=4).alignment = money_align
            ws.cell(row=row, column=4).border = thin_border

            ws.cell(row=row, column=5, value=round(group['total_balance'], 2)).font = total_font
            ws.cell(row=row, column=5).fill = total_fill
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=5).alignment = money_align
            ws.cell(row=row, column=5).border = thin_border

            row += 1  # blank row between groups

        # --- Grand Total ---
        grand = data['grand_total']
        ws.cell(row=row, column=1, value='').font = grand_font
        ws.cell(row=row, column=1).fill = grand_fill
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value='Grand Total').font = grand_font
        ws.cell(row=row, column=2).fill = grand_fill
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=round(grand['debit'], 2)).font = grand_font
        ws.cell(row=row, column=3).fill = grand_fill
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=3).alignment = money_align
        ws.cell(row=row, column=3).border = thin_border

        ws.cell(row=row, column=4, value=round(grand['credit'], 2)).font = grand_font
        ws.cell(row=row, column=4).fill = grand_fill
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=4).alignment = money_align
        ws.cell(row=row, column=4).border = thin_border

        ws.cell(row=row, column=5, value=round(grand['balance'], 2)).font = grand_font
        ws.cell(row=row, column=5).fill = grand_fill
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=5).alignment = money_align
        ws.cell(row=row, column=5).border = thin_border

        # --- Output ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = 'trial_balance_%s_%s.xlsx' % (wizard.date_from, wizard.date_to)
        return request.make_response(
            output.getvalue(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', 'attachment; filename=%s' % filename),
            ],
        )