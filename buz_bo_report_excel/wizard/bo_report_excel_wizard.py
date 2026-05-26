import base64
import io
import logging

from odoo import api, fields, models, _
from odoo.exceptions import UserError

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
except ImportError:
    openpyxl = None


class BOReportExcelWizard(models.TransientModel):
    _name = "bo.report.excel.wizard"
    _description = "Export Purchase Blanket Orders to Excel"

    date_from = fields.Date(string="Date From")
    date_to = fields.Date(string="Date To")
    vendor_id = fields.Many2one(
        "res.partner", string="Vendor",
        domain="[('supplier_rank', '>', 0)]",
    )
    state = fields.Selection([
        ("draft", "Draft"),
        ("ongoing", "Ongoing"),
        ("in_progress", "Confirmed"),
        ("open", "Bid Selection"),
        ("done", "Closed"),
        ("cancel", "Cancelled"),
    ], string="Status")
    requisition_ids = fields.Many2many(
        "purchase.requisition", string="Purchase Agreements",
        help="Pre-selected records from list view",
    )

    # ── Helpers ─────────────────────────────────────────────────
    def _get_requisition_domain(self):
        domain = []
        if self.requisition_ids:
            domain.append(("id", "in", self.requisition_ids.ids))
        else:
            if self.date_from:
                domain.append(("ordering_date", ">=", self.date_from))
            if self.date_to:
                domain.append(("ordering_date", "<=", self.date_to))
            if self.vendor_id:
                domain.append(("vendor_id", "=", self.vendor_id.id))
            if self.state:
                domain.append(("state", "=", self.state))
        return domain

    @staticmethod
    def _number_to_words(value):
        units = [
            "Zero", "One", "Two", "Three", "Four", "Five", "Six",
            "Seven", "Eight", "Nine", "Ten", "Eleven", "Twelve",
            "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen",
            "Eighteen", "Nineteen",
        ]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty",
                "Sixty", "Seventy", "Eighty", "Ninety"]

        def _sub(n):
            words = []
            if n >= 100:
                words += [units[n // 100], "Hundred"]
                n %= 100
            if n >= 20:
                words.append(tens[n // 10])
                if n % 10:
                    words.append(units[n % 10])
            elif n > 0:
                words.append(units[n])
            return " ".join(words)

        if value == 0:
            return units[0]
        if value < 0:
            return "Minus " + BOReportExcelWizard._number_to_words(abs(value))
        parts = []
        for div, name in [(1_000_000_000, "Billion"),
                          (1_000_000, "Million"),
                          (1_000, "Thousand"),
                          (1, "")]:
            q, value = divmod(value, div)
            if q:
                parts.append(f"{_sub(q)} {name}".strip())
        return " ".join(parts)

    # ── Styles (single source of truth) ─────────────────────────
    @staticmethod
    def _make_styles():
        """Return a dict of reusable openpyxl style objects."""
        navy = "1F4E79"
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        return {
            "border": border,
            "border_thick": Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="medium"), bottom=Side(style="thin"),
            ),
            "border_bottom_medium": Border(bottom=Side(style="medium")),
            "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
            "right": Alignment(horizontal="right", vertical="center", wrap_text=True),
            "font_title": Font(name="Calibri", bold=True, size=16, color=navy),
            "font_company": Font(name="Calibri", bold=True, size=12, color=navy),
            "font_company_info": Font(name="Calibri", size=9, color="555555"),
            "font_doc_title_th": Font(name="Calibri", bold=True, size=18, color=navy),
            "font_doc_title_en": Font(name="Calibri", bold=True, size=13, color=navy),
            "font_label": Font(name="Calibri", bold=True, size=10, color=navy),
            "font_value": Font(name="Calibri", size=10),
            "font_header": Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
            "font_data": Font(name="Calibri", size=10),
            "font_total_label": Font(name="Calibri", bold=True, size=11, color=navy),
            "font_total_value": Font(name="Calibri", bold=True, size=11, color=navy),
            "font_amount_text": Font(name="Calibri", italic=True, size=9),
            "font_sign_label": Font(name="Calibri", bold=True, size=9, color=navy),
            "font_sign_sub": Font(name="Calibri", size=8, color="888888"),
            "font_condition": Font(name="Calibri", size=8, color="444444"),
            "font_condition_title": Font(name="Calibri", bold=True, size=9, color=navy),
            "fill_header": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
            "fill_alt_row": PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid"),
            "fill_total": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            "fill_sign_box": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
            "num_2d": "#,##0.00",
            "num_4d": "#,##0.0000",
        }

    # ── Cell helper ─────────────────────────────────────────────
    @staticmethod
    def _cell(sheet, row, col, value, font=None, alignment=None, border=None, fill=None, number_format=None):
        cell = sheet.cell(row=row, column=col, value=value)
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if fill:
            cell.fill = fill
        if number_format:
            cell.number_format = number_format
        return cell

    # ── Main export ─────────────────────────────────────────────
    def action_export_excel(self):
        if openpyxl is None:
            raise UserError(_("openpyxl library is not installed."))

        requisitions = self.env["purchase.requisition"].search(
            self._get_requisition_domain(), order="ordering_date desc, id desc",
        )
        if not requisitions:
            raise UserError(_("No Purchase Agreements found with the selected criteria."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Purchase Order"
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        S = self._make_styles()
        C = lambda r, c, v, **kw: self._cell(ws, r, c, v, **kw)  # shorthand

        # ── Column widths (A-K, 11 cols) ───────────────────────
        col_widths = [6, 16, 14, 34, 16, 14, 10, 10, 14, 10, 16]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        company = self.env.company
        req = requisitions[0]
        vendor = req.vendor_id or self.env["res.partner"]

        # ── Header section ──────────────────────────────────────
        # Row 1: Company name
        ws.merge_cells("A1:K1")
        C(1, 1, company.name or "", font=S["font_company"], alignment=S["left"])
        ws.row_dimensions[1].height = 22

        # Row 2: Company address
        addr = " ".join(p for p in [company.street, company.street2] if p)
        ws.merge_cells("A2:K2")
        C(2, 1, addr, font=S["font_company_info"], alignment=S["left"])

        # Row 3: Phone / Fax / Email
        ws.merge_cells("A3:K3")
        contact_parts = []
        company_phone = getattr(company, 'phone', '')
        company_fax = getattr(company, 'fax', '')
        company_email = getattr(company, 'email', '')
        if company_phone:
            contact_parts.append(f"โทร. {company_phone}")
        if company_fax:
            contact_parts.append(f"แฟกซ์ {company_fax}")
        if company_email:
            contact_parts.append(f"E-mail : {company_email}")
        C(3, 1, "   ".join(contact_parts), font=S["font_company_info"], alignment=S["left"])

        # Row 4: Tax ID
        ws.merge_cells("A4:K4")
        C(4, 1, f"เลขที่ผู้เสียภาษี : {company.vat or ''}", font=S["font_company_info"], alignment=S["left"])

        # Row 5: separator line
        ws.row_dimensions[5].height = 8
        for col in range(1, 12):
            C(5, col, "", border=S["border_bottom_medium"])

        # Row 6-7: Document title (2 lines, centered)
        ws.row_dimensions[6].height = 30
        ws.merge_cells("A6:K6")
        C(6, 1, "ใบสั่งซื้อ", font=S["font_doc_title_th"], alignment=S["center"])

        ws.merge_cells("A7:K7")
        C(7, 1, "PURCHASE ORDER", font=S["font_doc_title_en"], alignment=S["center"])

        date_issue = req.ordering_date.strftime("%b.%d,%Y").upper() if req.ordering_date else ""
        vendor_addr_lines = [p for p in [vendor.street, vendor.street2] if p]
        vendor_city_parts = [p for p in [vendor.city, getattr(vendor.state_id, "name", ""), vendor.zip, getattr(vendor.country_id, "name", "")] if p]
        vendor_line1 = vendor_addr_lines[0] if vendor_addr_lines else ""
        vendor_line2 = ", ".join(vendor_addr_lines[1:] + vendor_city_parts)
        delivery_location = "Mogen(Thailand)Co.,Ltd. 9 Moo 12 T.Beungkhumproy A.Lumlukka Pathum Thani 12150"
        contact_person = getattr(vendor, "contact_name", False) or vendor.name or ""
        payment_term = ""
        if hasattr(vendor, "property_supplier_payment_term_id") and vendor.property_supplier_payment_term_id:
            payment_term = vendor.property_supplier_payment_term_id.name or ""
        criteria = req.type_id.name or ""
        ref_number = req.name or ""
        ref_origin = req.origin or ""

        C(8, 1, "PO Number :", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("B8:D8")
        C(8, 2, ref_number, font=S["font_value"], alignment=S["left"], border=S["border"])
        C(8, 5, "Date issue :", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("F8:G8")
        C(8, 6, date_issue, font=S["font_value"], alignment=S["left"], border=S["border"])
        ws.merge_cells("H8:K8")
        C(8, 8, "", border=S["border"])

        ws.merge_cells("A9:G9")
        C(9, 1, "กรุณาแนบใบสั่งซื้อทุกครั้งที่ส่งสินค้าและวางบิล", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("H9:K9")
        C(9, 8, "Confirm order from supplier", font=S["font_label"], alignment=S["left"], border=S["border"])

        C(10, 1, "นามผู้ขาย / Vendor :", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("B10:D10")
        C(10, 2, vendor.name or "", font=S["font_value"], alignment=S["left"], border=S["border"])
        C(10, 5, "สถานที่ส่งสินค้า / Place of delivery :", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("F10:K10")
        C(10, 6, delivery_location, font=S["font_value"], alignment=S["left"], border=S["border"])

        C(11, 1, "ที่อยู่/ADDRESS :", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("B11:D11")
        C(11, 2, vendor_line1 or "", font=S["font_value"], alignment=S["left"], border=S["border"])
        C(11, 5, "Date of :", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("F11:K11")
        C(11, 6, req.schedule_date.strftime("%d-%b-%y") if req.schedule_date else "", font=S["font_value"], alignment=S["left"], border=S["border"])

        ws.merge_cells("A12:D12")
        C(12, 1, vendor_line2 or "", font=S["font_value"], alignment=S["left"], border=S["border"])
        C(12, 5, "Contact Person", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("F12:G12")
        C(12, 6, contact_person, font=S["font_value"], alignment=S["left"], border=S["border"])
        C(12, 8, "Payment term", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("I12:K12")
        C(12, 9, payment_term, font=S["font_value"], alignment=S["left"], border=S["border"])

        C(13, 1, "Criteria", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("B13:D13")
        C(13, 2, criteria, font=S["font_value"], alignment=S["left"], border=S["border"])
        C(13, 5, "Ref.Number", font=S["font_label"], alignment=S["left"], border=S["border"])
        ws.merge_cells("F13:K13")
        C(13, 6, ref_origin, font=S["font_value"], alignment=S["left"], border=S["border"])

        # Row 14: empty separator
        ws.row_dimensions[14].height = 8

        # ── Table header (row 16) ───────────────────────────────
        HEADER_ROW = 16
        headers_full = [
            "ITEM", "ITEM NO.", "IMAGE", "DESCRIPTION", "Supplier Name",
            "Date of Receipt", "Quantity", "Unit", "Price(USD)/Unit",
            "DISC.", "AMOUNT",
        ]
        ws.row_dimensions[HEADER_ROW].height = 28
        for col_idx, header in enumerate(headers_full, start=1):
            C(HEADER_ROW, col_idx, header,
              font=S["font_header"], fill=S["fill_header"],
              alignment=S["center"], border=S["border"])

        # ── Data rows ───────────────────────────────────────────
        DATA_START = 17
        row = DATA_START
        line_no = 1
        total_qty = 0.0
        total_amount = 0.0

        for req in requisitions:
            vendor_name = (req.vendor_id or vendor).name or ""
            for line in req.line_ids:
                qty = line.product_qty or 0.0
                price = line.price_unit or 0.0
                discount = getattr(line, "discount", 0.0) or 0.0
                amount = qty * price * (1 - discount / 100.0)
                total_qty += qty
                total_amount += amount

                item_code = getattr(line.product_id, "default_code", "") or ""
                desc = line.product_id.display_name or ""
                if line.product_description_variants:
                    desc += f"\n{line.product_description_variants}"
                date_str = req.schedule_date.strftime("%d/%m/%Y") if req.schedule_date else ""

                row_data = [
                    line_no,       # A: item no
                    item_code,     # B: item code
                    "",            # C: image (placeholder, image inserted below)
                    desc,          # D: description
                    vendor_name,   # E: supplier
                    date_str,      # F: date
                    qty,           # G: qty
                    line.product_uom_id.name or "",  # H: unit
                    price,         # I: price/unit
                    discount,      # J: disc %
                    amount,        # K: amount
                ]
                row_fill = S["fill_alt_row"] if line_no % 2 == 0 else None
                ws.row_dimensions[row].height = 55

                for col_idx, val in enumerate(row_data, start=1):
                    nf = None
                    al = S["center"] if col_idx in (1, 6, 7, 8) else S["left"]
                    if col_idx == 2:
                        al = S["left"]
                    elif col_idx == 3:
                        al = S["center"]  # image cell
                    elif col_idx == 4:
                        al = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    elif col_idx in (7, 9, 11):
                        nf = S["num_2d"]
                        al = S["right"]
                    elif col_idx == 10:
                        nf = S["num_2d"]
                        al = S["center"]
                    C(row, col_idx, val, font=S["font_data"],
                      alignment=al, border=S["border"],
                      fill=row_fill, number_format=nf)

                # ── Insert product image in IMAGE column (C) ──
                product = line.product_id
                img_field = product.image_1920 or product.image_128
                if img_field:
                    try:
                        from PIL import Image as PILImage

                        img_bytes = base64.b64decode(img_field)
                        pil_img = PILImage.open(io.BytesIO(img_bytes))

                        # Resize to fit image cell C and row height
                        max_w, max_h = 90, 55
                        w_ratio = max_w / pil_img.width
                        h_ratio = max_h / pil_img.height
                        ratio = min(w_ratio, h_ratio, 1.0)
                        new_w = int(pil_img.width * ratio)
                        new_h = int(pil_img.height * ratio)
                        pil_img = pil_img.resize((new_w, new_h), PILImage.LANCZOS)

                        resized_buf = io.BytesIO()
                        pil_img.save(resized_buf, format="PNG")
                        resized_buf.seek(0)

                        xl_img = XlImage(resized_buf)
                        xl_img.width = new_w
                        xl_img.height = new_h

                        # Center image in column C using OneCellAnchor
                        col_letter = "C"
                        char_w = ws.column_dimensions[col_letter].width or 14
                        col_px = char_w * 7 + 5  # approximate pixel width (Calibri 10pt)
                        offset_px = max(0, int((col_px - new_w) / 2))

                        anchor = OneCellAnchor()
                        anchor._from = AnchorMarker(
                            col=ws[f"{col_letter}{row}"].column - 1,
                            colOff=pixels_to_EMU(offset_px),
                            row=row - 1,
                            rowOff=pixels_to_EMU(3),
                        )
                        anchor.ext = XDRPositiveSize2D(
                            cx=pixels_to_EMU(new_w),
                            cy=pixels_to_EMU(new_h),
                        )
                        xl_img.anchor = anchor
                        ws.add_image(xl_img)
                    except Exception:
                        logger.debug("Failed to embed image for product %s", product.id, exc_info=True)
                row += 1
                line_no += 1

            if not req.line_ids:
                row += 1

        # ── Summary section ─────────────────────────────────────
        LAST_DATA = row - 1
        sep_row = row + 1  # blank row

        # Gross
        r = sep_row + 1
        ws.merge_cells(f"A{r}:H{r}")
        C(r, 1, "REMARK : First Lot สินค้าใหม่", font=S["font_label"], alignment=S["left"])
        ws.merge_cells(f"I{r}:J{r}")
        C(r, 9, "Gross Amount (USD)", font=S["font_total_label"], alignment=S["right"])
        C(r, 11, total_amount, font=S["font_total_value"], alignment=S["right"],
          number_format=S["num_2d"])

        # Discount
        r += 1
        ws.merge_cells(f"A{r}:H{r}")
        C(r, 1, "", font=S["font_value"], alignment=S["left"])
        ws.merge_cells(f"I{r}:J{r}")
        C(r, 9, "Discount ( USD)", font=S["font_label"], alignment=S["right"])
        C(r, 11, 0.0, font=S["font_value"], alignment=S["right"], number_format=S["num_2d"])

        # Balance
        r += 1
        ws.merge_cells(f"A{r}:H{r}")
        C(r, 1, "", font=S["font_value"], alignment=S["left"])
        ws.merge_cells(f"I{r}:J{r}")
        C(r, 9, "Balance (USD)", font=S["font_label"], alignment=S["right"])
        C(r, 11, total_amount, font=S["font_value"], alignment=S["right"],
          number_format=S["num_2d"])

        # Total
        r += 1
        ws.merge_cells(f"A{r}:H{r}")
        C(r, 1, "", font=S["font_value"])
        ws.merge_cells(f"I{r}:J{r}")
        C(r, 9, "TOTAL AMOUNT", font=S["font_total_label"], alignment=S["right"],
          fill=S["fill_total"])
        C(r, 11, total_amount, font=S["font_total_value"], alignment=S["right"],
          number_format=S["num_2d"], fill=S["fill_total"])

        # Total amount in words and final total row
        r += 1
        ws.merge_cells(f"A{r}:D{r}")
        C(r, 1, "Total Amount (USD)", font=S["font_label"], alignment=S["left"])
        ws.merge_cells(f"E{r}:H{r}")
        amount_text = f"{self._number_to_words(int(round(total_amount)))} US Dollar"
        C(r, 5, amount_text, font=S["font_amount_text"], alignment=S["left"])
        C(r, 10, "Total Amount (USD)", font=S["font_label"], alignment=S["right"])
        C(r, 11, total_amount, font=S["font_total_value"], alignment=S["right"], number_format=S["num_2d"])

        # ── Signature section ───────────────────────────────────
        r += 2
        sign_cols = [
            (1, 3, "ผู้สั่งซื้อ / PURCHASED BY", "วันที่ / Date"),
            (4, 7, "ผู้ตรวจสอบ / CHECKED BY", "วันที่ / Date"),
            (8, 11, "ผู้อนุมัติ / APPROVED BY", "วันที่ / Date"),
        ]
        for c_start, c_end, title, subtitle in sign_cols:
            ws.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
            C(r, c_start, title, font=S["font_sign_label"], alignment=S["center"],
              fill=S["fill_sign_box"], border=S["border"])
            for c in range(c_start + 1, c_end + 1):
                ws.cell(row=r, column=c).fill = S["fill_sign_box"]
                ws.cell(row=r, column=c).border = S["border"]
            # Signature space
            for blank in range(1, 4):
                ws.row_dimensions[r + blank].height = 18
                for c in range(c_start, c_end + 1):
                    ws.cell(row=r + blank, column=c).border = S["border"]
            # Date row
            ws.merge_cells(start_row=r + 4, start_column=c_start, end_row=r + 4, end_column=c_end)
            C(r + 4, c_start, subtitle, font=S["font_sign_sub"], alignment=S["center"],
              border=S["border"])
            for c in range(c_start + 1, c_end + 1):
                ws.cell(row=r + 4, column=c).border = S["border"]

        # ── Conditions section ──────────────────────────────────
        r += 7
        ws.merge_cells(f"A{r}:F{r}")
        C(r, 1, "*** หมายเหตุ / เงื่อนไขการส่งมอบ ***",
          font=S["font_condition_title"], alignment=S["left"])
        ws.merge_cells(f"G{r}:K{r}")
        C(r, 7, "*** REMARK / CONDITIONS ***",
          font=S["font_condition_title"], alignment=S["left"])

        conditions = [
            ("1. สินค้าที่ส่งมอบต้องมีคุณภาพ ขนาด ปริมาณ และจัดส่งตามกำหนด",
             "1. ALL SUPPLIES ARE TO BE MADE STRICTLY IN ACCORDANCE WITH OUR SPECIFICATION"),
            ("2. ถ้าไม่เป็นไปตามเงื่อนไขข้างต้น ผู้ซื้อมีสิทธิยกเลิกใบสั่งซื้อ",
             "2. THIS ORDER SHALL BE CANCELLED IF DELIVERY IS NOT MADE WITH TIME STATED"),
            ("3. ใบส่งของทุกฉบับต้องมีสำเนา 2 ฉบับ พร้อมระบุเลขที่ใบสั่งซื้อ",
             "3. ALL SUPPLIES SHOULD BE COVERED BY YOUR INVOICE IN TRIPLICATION"),
            ("4. กรรมสิทธิ์และความรับผิดชอบของสินค้าจะเปลี่ยนจากของท่านเป็นของบริษัทฯ",
             "4. TITLE AND RISK OF THE GOODS SHALL PASS FROM YOU TO US"),
            ("5. บริษัทฯ จะไม่รับผิดชอบในการชำระเงินสำหรับสินค้าที่ผู้ขายส่งเกิน",
             "5. INVOICE/DELIVERY NOTE MUST BE STATED THE NUMBER OF THIS ORDER"),
        ]
        for idx, (th, en) in enumerate(conditions, 1):
            cr = r + idx
            ws.merge_cells(f"A{cr}:F{cr}")
            C(cr, 1, th, font=S["font_condition"], alignment=S["left"])
            ws.merge_cells(f"G{cr}:K{cr}")
            C(cr, 7, en, font=S["font_condition"], alignment=S["left"])

        # ── Footer line ─────────────────────────────────────────
        fr = r + len(conditions) + 2
        ws.merge_cells(f"A{fr}:F{fr}")
        C(fr, 1, f"อนุมัติใช้วันที่ {req.ordering_date.strftime('%d/%m/%Y') if req.ordering_date else ''}",
          font=Font(size=7, color="999999"), alignment=S["left"])
        ws.merge_cells(f"G{fr}:K{fr}")
        C(fr, 7, "FM-PUR-02.REV.02",
          font=Font(size=7, color="999999"), alignment=S["right"])

        # ── Print setup ─────────────────────────────────────────
        ws.freeze_panes = f"A{DATA_START}"
        if LAST_DATA >= DATA_START:
            ws.auto_filter.ref = f"A{HEADER_ROW}:K{LAST_DATA}"

        ws.print_area = f"A1:K{fr}"
        ws.page_margins = openpyxl.worksheet.page.PageMargins(
            left=0.4, right=0.4, top=0.4, bottom=0.4)

        # ── Save & download ─────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        attachment = self.env["ir.attachment"].create({
            "name": f"PO_{req.name or 'report'}.xlsx",
            "datas": base64.b64encode(buf.read()),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }
