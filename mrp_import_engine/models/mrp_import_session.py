# -*- coding: utf-8 -*-
from odoo import api, fields, models, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

class MrpImportSession(models.Model):
    _name = 'mrp.import.session'
    _description = 'MRP Import Session'
    _order = 'create_date desc'

    name = fields.Char(string='Reference', required=True, copy=False, readonly=True, default=lambda self: _('New'))
    import_type = fields.Selection([
        ('bom', 'Bill of Materials'),
        ('mo', 'Manufacturing Order')
    ], string='Import Type', required=True)
    file = fields.Binary(string='File')
    filename = fields.Char(string='File Name')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('done', 'Done'),
        ('failed', 'Failed')
    ], string='Status', default='draft', readonly=True)
    
    total_records = fields.Integer(string='Total Records', default=0, readonly=True)
    success_records = fields.Integer(string='Success Records', default=0, readonly=True)
    failed_records = fields.Integer(string='Failed Records', default=0, readonly=True)
    
    log_ids = fields.One2many('mrp.import.log', 'session_id', string='Logs')
    batch_id = fields.Char(string='Batch ID', readonly=True)
    external_ref = fields.Char(string='External Ref')

    bom_staging_ids = fields.One2many('mrp.bom.staging', 'session_id', string='BOM Staging')
    mo_staging_ids = fields.One2many('mrp.mo.staging', 'session_id', string='MO Staging')

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', 'New') == 'New':
                vals['name'] = self.env['ir.sequence'].next_by_code('mrp.import.session') or 'New'
        return super().create(vals_list)

    def action_process(self):
        for record in self:
            if record.state != 'draft':
                continue
            record.state = 'processing'
            if record.file:
                record._parse_file()
            
            if record.import_type == 'bom':
                record.process_bom_import()
            else:
                record.process_mo_import()

    def _map_bom_type_local(self, val):
        if not val:
            return 'normal'
        val_str = str(val).strip().lower()
        if val_str in ['phantom', 'kit', 'ชุดประกอบ']:
            return 'phantom'
        return 'normal'

    def _parse_file(self):
        self.ensure_one()
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("The 'openpyxl' python library is required to parse Excel files. Please install it."))
            
        try:
            file_data = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_data), data_only=True)
            
            if self.import_type == 'bom':
                # Parse BOM
                ws_bom = wb['BOM'] if 'BOM' in wb.sheetnames else wb.worksheets[0]
                boms_data = {}
                # Start at row 3 (idx 2) due to title and hints
                for row in ws_bom.iter_rows(min_row=3, values_only=True):
                    if not row or not row[0]: continue
                    bom_code = str(row[0]).strip()
                    boms_data[bom_code] = {
                        'session_id': self.id,
                        'bom_code': bom_code,
                        'product_default_code': str(row[1]) if row[1] else '',
                        'product_qty': float(row[2]) if row[2] else 1.0,
                        'bom_type': str(row[3]) if len(row) > 3 and row[3] else 'Manufacture this product',
                        'version': str(row[4]) if len(row) > 4 and row[4] else '',
                        'lines': []
                    }
                
                # Parse BOM Lines
                if 'BOM Lines' in wb.sheetnames:
                    ws_line = wb['BOM Lines']
                    for row in ws_line.iter_rows(min_row=3, values_only=True):
                        if not row or not row[0]: continue
                        bom_code = str(row[0]).strip()
                        if bom_code in boms_data:
                            boms_data[bom_code]['lines'].append({
                                'component_default_code': str(row[1]).strip() if row[1] else '',
                                'quantity': float(row[2]) if row[2] else 1.0,
                                'uom': str(row[3]).strip() if len(row) > 3 and row[3] else '',
                                'operation': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                            })
                
                # Create Stagings
                for bom_code, values in boms_data.items():
                    staging = self.env['mrp.bom.staging'].create({
                        'session_id': self.id,
                        'bom_code': values['bom_code'],
                        'product_default_code': values['product_default_code'],
                        'product_qty': values['product_qty'],
                        'version': values['version'],
                        'bom_type': self._map_bom_type_local(values.get('bom_type', ''))
                    })
                    for l in values['lines']:
                        self.env['mrp.bom.line.staging'].create({
                            'bom_staging_id': staging.id,
                            'bom_code': bom_code,
                            'component_default_code': l['component_default_code'],
                            'quantity': l['quantity'],
                            'uom': l['uom'],
                            'operation': l['operation'],
                        })

            else:
                # Parse MO
                ws_mo = wb['Manufacturing Orders'] if 'Manufacturing Orders' in wb.sheetnames else wb.worksheets[0]
                for row in ws_mo.iter_rows(min_row=3, values_only=True):
                    if not row or not row[0]: continue
                    self.env['mrp.mo.staging'].create({
                        'session_id': self.id,
                        'mo_name': str(row[0]),
                        'product_default_code': str(row[1]) if row[1] else '',
                        'product_qty': float(row[2]) if row[2] else 1.0,
                        'bom_code': str(row[3]) if len(row) > 3 and row[3] else '',
                        'bom_type': self._map_bom_type_local(str(row[4]) if len(row) > 4 and row[4] else 'Manufacture this product'),
                        'operation_type': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                        'work_center': str(row[6]).strip() if len(row) > 6 and row[6] else '',
                        'planned_start_date': row[7] if len(row) > 7 and row[7] else False,
                        'planned_finish_date': row[8] if len(row) > 8 and row[8] else False,
                    })
                    
        except Exception as e:
            _logger.exception("Failed to parse import file: %s", str(e))
            self.state = 'failed'
            self.env['mrp.import.log'].create({
                'session_id': self.id,
                'row_number': 0,
                'state': 'failed',
                'message': f"Failed to parse file: {str(e)}"
            })
            raise UserError(_("Failed to parse file: %s") % str(e))

    def process_bom_import(self):
        # Delegate to service
        self.env['mrp.import.bom.service'].process_session(self.id)

    def process_mo_import(self):
        # Delegate to service
        self.env['mrp.import.mo.service'].process_session(self.id)
