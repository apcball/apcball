# -*- coding: utf-8 -*-
"""
Landed Cost Distribution Report — Excel Export

เรียกจาก action menu บน stock.landed.cost form/list view
แสดง additional cost แตกหารตาม cost line แบบ matrix

ตัวอย่างผลลัพธ์:
| Product Code | Product | Qty | Former Cost | Freight | Duty | Insurance | Total Additional | Final Cost | LC/Unit |
| SKU-001      | สินค้าA | 50  | 5,000       | 625     | 312  | 150       | 1,087            | 6,087      | 21.74   |
"""

import logging
import os
from collections import defaultdict

from odoo import models, fields, api, _
from odoo.exceptions import UserError
from odoo.tools import float_round

_logger = logging.getLogger(__name__)

COST_LINE_PREDEFINED = 'product'


class StockLandedCost(models.Model):
    _inherit = 'stock.landed.cost'

    def action_export_distribution_excel(self):
        """
        Export landed cost distribution report for selected LC document(s).
        Called from action menu on stock.landed.cost.
        """
        landed_costs = self.filtered(lambda lc: lc.state == 'done')
        if not landed_costs:
            raise UserError(_('Please select validated Landed Cost document(s).'))

        lc_loc_model = self.env['stock.valuation.layer.landed.cost']

        sheets = []

        for lc in landed_costs:
            sheet_name = ('LC_%s' % (lc.name or str(lc.id)))[:31].replace('/', '_').replace('\\', '_')

            cost_lines = lc.cost_lines
            cost_line_columns = self._get_cost_line_report_columns(lc)
            adj_lines = lc.valuation_adjustment_lines.filtered(lambda l: l.move_id)

            # ── Recompute per-cost-line allocation per product ──
            # We need to split each cost_line's amount to each product using its split_method
            product_cost_breakdown = self._compute_per_cost_line_allocation(cost_lines, adj_lines)

            # ── Section 1: LC Info + Cost Lines Summary ──
            rows = [
                ['Landed Cost Document', lc.name or ''],
                ['Date', str(lc.date or (lc.create_date.date() if lc.create_date else ''))],
                ['', ''],
                ['Additional Cost Lines', '', '', '', ''],
                ['No.', 'Description', 'Split Method', 'Product', 'Amount'],
            ]

            for idx, col in enumerate(cost_line_columns, 1):
                rows.append([
                    idx,
                    col['name'],
                    col['split_method'],
                    col['product_name'],
                    col['amount'],
                ])

            total_lc_amount = sum(cl.price_unit for cl in cost_lines)
            rows.append(['', '', '', 'TOTAL', total_lc_amount])
            rows.append(['', ''])

            # ── Section 2: Cost Allocation Matrix (MAIN) ──
            # Build header: Product Code | Product | Qty | Former Cost | CostLine1 | CostLine2 | ... | Total Additional | Final Cost | LC/Unit
            cl_columns = []
            for col in cost_line_columns:
                cl_columns.append(col['name'])

            matrix_header = ['Product Code', 'Product', 'Qty', 'Former Cost']
            matrix_header.extend(cl_columns)
            matrix_header.extend(['Total Additional Cost', 'Final Cost', 'LC/Unit'])

            rows.append(['Cost Allocation by Product', '', '', '', ''])
            rows.append(matrix_header)

            # Group adj_lines by move (one per receipt line, NOT one per cost line)
            seen_moves = set()
            grouped_adjs = []
            for adj in adj_lines:
                move_key = (adj.move_id.id, adj.product_id.id)
                if move_key not in seen_moves:
                    seen_moves.add(move_key)
                    grouped_adjs.append(adj)

            for adj in grouped_adjs:
                product = adj.product_id
                product_code = product.default_code or ''
                breakdown = product_cost_breakdown.get(adj.id, {})

                row_vals = [product_code, product.display_name, adj.quantity, adj.former_cost]

                # Per-cost-line values
                per_line_total = 0.0
                for col in cost_line_columns:
                    cl = col['cost_line']
                    cl_amount = breakdown.get(cl.id, 0.0) if cl else 0.0
                    row_vals.append(cl_amount)
                    per_line_total += cl_amount

                lc_per_unit = (per_line_total / adj.quantity) if adj.quantity else 0
                row_vals.append(per_line_total)
                row_vals.append(adj.former_cost + per_line_total)
                row_vals.append(lc_per_unit)
                rows.append(row_vals)

            # Total row
            total_row = ['', 'TOTAL', '', '']
            grand_total_additional = 0.0
            for col in cost_line_columns:
                cl = col['cost_line']
                cl_total = sum(
                    product_cost_breakdown.get(adj.id, {}).get(cl.id, 0.0) if cl else 0.0
                    for adj in grouped_adjs
                )
                total_row.append(cl_total)
                grand_total_additional += cl_total

            total_row.append(grand_total_additional)
            total_row.append(sum(adj.former_cost for adj in grouped_adjs) + grand_total_additional)
            total_row.append('')
            rows.append(total_row)
            rows.append(['', ''])

            # ── Section 3: Warehouse Distribution ──
            rows.append(['Warehouse Distribution', '', '', '', ''])
            rows.append(['Warehouse', 'Product', 'Qty', 'LC Value', 'LC/Unit'])

            lc_records = lc_loc_model.search([
                ('landed_cost_id', '=', lc.id),
            ])

            wh_total_lc = 0.0
            for lc_rec in lc_records:
                layer = lc_rec.valuation_layer_id
                product = layer.product_id if layer else None
                rows.append([
                    lc_rec.warehouse_id.display_name if lc_rec.warehouse_id else 'N/A',
                    product.default_code or (product.display_name if product else 'N/A'),
                    lc_rec.quantity,
                    lc_rec.landed_cost_value,
                    lc_rec.unit_landed_cost,
                ])
                wh_total_lc += lc_rec.landed_cost_value

            rows.append(['', '', 'TOTAL', wh_total_lc, ''])

            sheets.append({'sheet': sheet_name, 'rows': rows})

            all_sheets = sheets

        filename = 'landed_cost_distribution_%s.xlsx' % fields.Date.today().strftime('%Y%m%d')
        filepath = '/tmp/%s' % filename

        self._write_excel(filepath, all_sheets)

        return {
            'type': 'ir.actions.act_url',
            'url': '/stock_fifo_by_location/download_report?filepath=%s&filename=%s' % (filepath, filename),
            'target': 'self',
        }

    def _get_cost_line_report_columns(self, landed_cost):
        """
        Build report cost columns from the actual additional costs first, then
        append active products marked as landed-cost products.

        Odoo stores the checkbox as product.template.landed_cost_ok. Including
        those products here makes the report layout stable, while allocations
        still come from the landed cost document's real cost lines.
        """
        columns = []
        included_product_ids = set()

        for cost_line in landed_cost.cost_lines:
            product = cost_line.product_id
            if product:
                included_product_ids.add(product.id)
            columns.append({
                'cost_line': cost_line,
                'name': cost_line.name or (product.display_name if product else _('Cost')),
                'split_method': cost_line.split_method or '',
                'product_name': product.display_name if product else '',
                'amount': cost_line.price_unit,
            })

        if COST_LINE_PREDEFINED == 'product':
            Product = self.env['product.product']
            if 'landed_cost_ok' not in Product._fields:
                return columns

            domain = [('landed_cost_ok', '=', True)]
            if 'active' in Product._fields:
                domain.append(('active', '=', True))
            if 'company_id' in Product._fields:
                domain.append(('company_id', 'in', [False, landed_cost.company_id.id]))

            landed_cost_products = Product.search(domain, order='default_code, name')
            for product in landed_cost_products:
                if product.id in included_product_ids:
                    continue
                columns.append({
                    'cost_line': False,
                    'name': product.display_name,
                    'split_method': '',
                    'product_name': product.display_name,
                    'amount': 0.0,
                })

        return columns

    def _compute_per_cost_line_allocation(self, cost_lines, adj_lines):
        """
        Recompute per-cost-line allocation for each product.

        Returns: {adj_line_id: {cost_line_id: amount}}

        Note: Odoo creates one valuation_adjustment_line per (move, cost_line).
        So for 2 products × 4 cost lines = 8 adj lines.
        We group by move and sum cost_line contributions.
        """
        result = {}

        # Group adj_lines by move_id — collect all adj lines for same product
        move_groups = defaultdict(list)
        for adj in adj_lines:
            move_groups[(adj.move_id.id, adj.product_id.id)].append(adj)

        for (move_id, product_id), adjs_for_product in move_groups.items():
            # Use first adj as representative for qty/former_cost
            first_adj = adjs_for_product[0]
            product = first_adj.product_id
            qty = first_adj.quantity or 0
            weight = product.weight * qty if product.weight else 0
            volume = product.volume * qty if product.volume else 0
            former_cost = first_adj.former_cost or 0

            # Sum all adj lines for this product to get total additional
            total_additional = sum(adj.additional_landed_cost for adj in adjs_for_product)

            for cl in cost_lines:
                split_method = cl.split_method or 'equal'
                price_unit = cl.price_unit or 0

                # Find adj line that matches this cost line
                # Odoo creates adj lines with former_cost per cost line — 
                # but we need to figure out which adj belongs to which cost line.
                # Strategy: since each adj has additional_landed_cost, and we know 
                # the split method, we can recompute the allocation directly.

                pass  # computed below in the loop

        # ── Direct recomputation approach ──
        # Build move data for split calculations — one per unique (move, product)
        moves_data = []
        seen = set()
        for adj in adj_lines:
            key = (adj.move_id.id, adj.product_id.id)
            if key in seen:
                continue
            seen.add(key)

            product = adj.product_id
            qty = adj.quantity or 0
            moves_data.append({
                'adj_id': adj.id,
                'product': product,
                'quantity': qty,
                'weight': product.weight * qty if product.weight else 0,
                'volume': product.volume * qty if product.volume else 0,
                'former_cost': adj.former_cost or 0,
            })

        for cl in cost_lines:
            split_method = cl.split_method or 'equal'
            price_unit = cl.price_unit or 0

            if split_method == 'equal':
                count = len(moves_data) or 1
                per_unit = price_unit / count
                for md in moves_data:
                    result.setdefault(md['adj_id'], {})[cl.id] = float_round(per_unit, precision_digits=2)

            elif split_method == 'by_quantity':
                total_qty = sum(md['quantity'] for md in moves_data) or 1
                for md in moves_data:
                    share = price_unit * (md['quantity'] / total_qty)
                    result.setdefault(md['adj_id'], {})[cl.id] = float_round(share, precision_digits=2)

            elif split_method == 'by_weight':
                total_weight = sum(md['weight'] for md in moves_data) or 1
                for md in moves_data:
                    share = price_unit * (md['weight'] / total_weight)
                    result.setdefault(md['adj_id'], {})[cl.id] = float_round(share, precision_digits=2)

            elif split_method == 'by_volume':
                total_volume = sum(md['volume'] for md in moves_data) or 1
                for md in moves_data:
                    share = price_unit * (md['volume'] / total_volume)
                    result.setdefault(md['adj_id'], {})[cl.id] = float_round(share, precision_digits=2)

            elif split_method == 'by_current_cost_price':
                total_cost = sum(md['former_cost'] for md in moves_data) or 1
                for md in moves_data:
                    share = price_unit * (md['former_cost'] / total_cost)
                    result.setdefault(md['adj_id'], {})[cl.id] = float_round(share, precision_digits=2)

            else:
                count = len(moves_data) or 1
                per_unit = price_unit / count
                for md in moves_data:
                    result.setdefault(md['adj_id'], {})[cl.id] = float_round(per_unit, precision_digits=2)

        return result

    def _write_excel(self, filepath, sheets):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

        wb = Workbook()
        default_sheet = wb.active

        header_font_white = Font(bold=True, size=10, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        light_blue_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        total_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        section_font = Font(bold=True, size=11, color='1F4E79')
        info_font = Font(bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # Detect header rows by first cell value
        HEADER_FIRST_CELLS = {
            'Product Code', 'Warehouse', 'No.',
        }

        for i, sheet_data in enumerate(sheets):
            if i == 0:
                ws = default_sheet
                ws.title = sheet_data['sheet'][:31]
            else:
                ws = wb.create_sheet(title=sheet_data['sheet'][:31])

            rows = sheet_data['rows']
            for row_idx, row in enumerate(rows):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)

                    first_val = str(row[0]) if row and row[0] else ''

                    # Section titles
                    if first_val in ('Additional Cost Lines', 'Cost Allocation by Product',
                                     'Warehouse Distribution'):
                        cell.font = section_font
                        continue

                    # Header rows (table headers)
                    if first_val in HEADER_FIRST_CELLS:
                        for c in range(1, len(row) + 1):
                            rc = ws.cell(row=row_idx + 1, column=c)
                            rc.font = header_font_white
                            rc.fill = header_fill
                            rc.alignment = Alignment(horizontal='center', wrap_text=True)
                            rc.border = thin_border
                        # Light blue background for cost line columns (between Former Cost and Total Additional)
                        if first_val == 'Product Code':
                            for c in range(5, len(row) - 2):
                                rc = ws.cell(row=row_idx + 1, column=c)
                                rc.fill = PatternFill(
                                    start_color='2E75B6', end_color='2E75B6', fill_type='solid'
                                )
                        continue

                    # Total row
                    if first_val == '' and len(row) > 1 and str(row[1]).upper() == 'TOTAL':
                        for c in range(1, len(row) + 1):
                            rc = ws.cell(row=row_idx + 1, column=c)
                            rc.font = total_font
                            rc.fill = total_fill
                            rc.border = thin_border
                        continue

                    # LC info rows
                    if first_val in ('Landed Cost Document', 'Date'):
                        cell.font = info_font
                        continue

                    # Number formatting
                    if isinstance(value, (int, float)) and not isinstance(value, bool):
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        wb.save(filepath)
        _logger.info("Landed Cost Distribution Report saved to %s", filepath)
