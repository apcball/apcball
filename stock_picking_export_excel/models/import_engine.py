import base64
import io
import openpyxl
from datetime import datetime
from odoo import models, api, _
from odoo.exceptions import UserError

class ImportEngine(models.AbstractModel):
    _name = 'import.engine'
    _description = 'Import Engine for Stock Picking'

    @api.model
    def import_excel_to_picking(self, picking, file_data):
        if not file_data:
            raise UserError(_("No file uploaded."))
            
        try:
            excel_data = base64.b64decode(file_data)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(excel_data), data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Invalid Excel file: %s" % str(e)))

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise UserError(_("Excel file is empty or missing headers."))

        headers = [str(x).strip().lower() if x else '' for x in header_row]
        
        try:
            product_code_idx = headers.index('product code')
            qty_idx = headers.index('qty')
        except ValueError:
            raise UserError(_("Missing required columns: 'Product Code' or 'Qty'. Please use the template exported from the system."))
            
        cost_idx = headers.index('cost') if 'cost' in headers else -1
        desc_idx = next((i for i, h in enumerate(headers) if 'description' in h or 'product name' in h), -1)
        
        # Prefer exact match for scheduled date, else look for just date
        date_idx = next((i for i, h in enumerate(headers) if 'scheduled date' in h), -1)
        if date_idx == -1:
             date_idx = next((i for i, h in enumerate(headers) if 'date' == h.strip()), -1)
             
        effective_date_idx = next((i for i, h in enumerate(headers) if 'effective date' in h), -1)

        moves_data = []
        picking_effective_date = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Pad row if it's shorter than header
            row = list(row) + [None] * (len(headers) - len(row))
            if not row[product_code_idx]:
                continue
                
            code = str(row[product_code_idx]).strip()
            qty = row[qty_idx] or 0.0
            cost = row[cost_idx] if cost_idx >= 0 and row[cost_idx] else 0.0
            
            try:
                qty = float(qty)
            except ValueError:
                raise UserError(_("Invalid quantity for product code %s: %s") % (code, qty))
                
            try:
                cost = float(cost)
            except ValueError:
                raise UserError(_("Invalid cost for product code %s: %s") % (code, cost))

            product = self.env['product.product'].search([('default_code', '=', code)], limit=1)
            if not product and 'old_product_code' in self.env['product.product']._fields:
                product = self.env['product.product'].search([('old_product_code', '=', code)], limit=1)
                
            if not product:
                raise UserError(_("Product not found in system with Product Code or Old Product Code: %s") % code)
                
            desc = str(row[desc_idx]).strip() if desc_idx >= 0 and row[desc_idx] else product.display_name
            date_val = row[date_idx] if date_idx >= 0 else None
            
            if effective_date_idx >= 0 and row[effective_date_idx] and not picking_effective_date:
                picking_effective_date = row[effective_date_idx]
                
            move_vals = {
                'name': desc,
                'description_picking': desc,
                'product_id': product.id,
                'product_uom': product.uom_id.id,
                'product_uom_qty': qty,
                'location_id': picking.location_id.id,
                'location_dest_id': picking.location_dest_id.id,
                'picking_id': picking.id,
                'company_id': picking.company_id.id,
                'price_unit': cost,
            }
            
            if date_val:
                try:
                    if isinstance(date_val, datetime):
                        move_vals['date'] = date_val
                        move_vals['date_deadline'] = date_val
                    elif isinstance(date_val, str):
                        # try parse simple format
                        parsed_date = datetime.strptime(date_val.split('.')[0], '%Y-%m-%d %H:%M:%S')
                        move_vals['date'] = parsed_date
                        move_vals['date_deadline'] = parsed_date
                except Exception:
                    pass # ignore wrong date formats or fallback to default

            if 'custom_cost_price' in self.env['stock.move']._fields:
                move_vals['custom_cost_price'] = cost
                
            moves_data.append(move_vals)
            
        if not moves_data:
            raise UserError(_("No valid data rows found in the file."))
            
        self.env['stock.move'].create(moves_data)
        
        if picking_effective_date:
            try:
                if isinstance(picking_effective_date, datetime):
                    picking.write({'date_done': picking_effective_date})
                elif isinstance(picking_effective_date, str):
                    parsed_eff_date = datetime.strptime(picking_effective_date.split('.')[0], '%Y-%m-%d %H:%M:%S')
                    picking.write({'date_done': parsed_eff_date})
            except Exception:
                pass
        
        return True
