import logging, os
logging.basicConfig(level=logging.INFO)

lc_model = env['stock.landed.cost']
done_lcs = lc_model.search([('state', '=', 'done')], limit=1)
if not done_lcs:
    print("No done landed costs found")
else:
    print("Found LC: %s (id=%d)" % (done_lcs.name or 'N/A', done_lcs.id))
    print("Cost lines: %d" % len(done_lcs.cost_lines))
    print("Adjustment lines: %d" % len(done_lcs.valuation_adjustment_lines))
    try:
        result = done_lcs.action_export_distribution_excel()
        print("Result type: %s" % result.get('type', 'unknown'))
        filepath = result['url'].split('filepath=')[1].split('&')[0]
        print("Filepath: %s" % filepath)
        if os.path.exists(filepath):
            print("File size: %d bytes" % os.path.getsize(filepath))
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            print("Sheets: %s" % str(wb.sheetnames))
            for sn in wb.sheetnames:
                ws = wb[sn]
                found_matrix = False
                for row in ws.iter_rows(values_only=True):
                    row_str = str(row)
                    if 'Matrix' in row_str:
                        found_matrix = True
                        break
                if found_matrix:
                    print("  MATRIX found in sheet: %s" % sn)
                else:
                    print("  No matrix in sheet: %s (rows=%d)" % (sn, ws.max_row))
        else:
            print("File not found")
    except Exception as e:
        print("ERROR: %s" % str(e))
        import traceback
        traceback.print_exc()
