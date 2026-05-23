import os
from openpyxl import load_workbook

lc_model = env['stock.landed.cost']
done_lcs = lc_model.search([('state', '=', 'done')], limit=3)
print("Found %d done LCs" % len(done_lcs))

for lc in done_lcs:
    print("\nLC: %s | Cost Lines: %d | Adj Lines: %d" % (
        lc.name, len(lc.cost_lines), len(lc.valuation_adjustment_lines.filtered(lambda l: l.move_id))
    ))
    for cl in lc.cost_lines:
        print("  Cost Line: %s | Split: %s | Amount: %.2f" % (cl.name, cl.split_method, cl.price_unit))

# Use the one with most cost lines for best demo
best_lc = max(done_lcs, key=lambda lc: len(lc.cost_lines))
print("\nExporting LC: %s (%d cost lines)" % (best_lc.name, len(best_lc.cost_lines)))

result = best_lc.action_export_distribution_excel()
filepath = result['url'].split('filepath=')[1].split('&')[0]
print("Filepath: %s" % filepath)

if os.path.exists(filepath):
    wb = load_workbook(filepath)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print("\n=== Sheet: %s (rows=%d) ===" % (sn, ws.max_row))
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [str(c) if c is not None else '' for c in row]
            print("Row %2d: %s" % (row_idx, ' | '.join(cells)))
