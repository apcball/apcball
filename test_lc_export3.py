import os
from openpyxl import load_workbook

# Create test data with multiple cost lines
product_model = env['product.product']
wh = env['stock.warehouse'].search([('company_id', '=', env.company.id)], limit=1)
supplier_loc = env.ref('stock.stock_location_suppliers')

# Create 2 test products
p1 = product_model.create({
    'name': 'Test LC Report Product A',
    'default_code': 'TEST-LC-A',
    'type': 'product',
    'standard_price': 200,
    'weight': 5.0,
    'categ_id': env.ref('product.product_category_all').id,
})
p2 = product_model.create({
    'name': 'Test LC Report Product B',
    'default_code': 'TEST-LC-B',
    'type': 'product',
    'standard_price': 250,
    'weight': 3.0,
    'categ_id': env.ref('product.product_category_all').id,
})

# Create a picking with 2 products
picking = env['stock.picking'].create({
    'picking_type_id': wh.in_type_id.id,
    'location_id': supplier_loc.id,
    'location_dest_id': wh.lot_stock_id.id,
})

for p, qty in [(p1, 10), (p2, 5)]:
    env['stock.move'].create({
        'name': 'Receipt %s' % p.default_code,
        'product_id': p.id,
        'product_uom_qty': qty,
        'product_uom': p.uom_id.id,
        'picking_id': picking.id,
        'location_id': supplier_loc.id,
        'location_dest_id': wh.lot_stock_id.id,
        'price_unit': p.standard_price,
    })

picking.action_confirm()
picking.action_assign()
for ml in picking.move_line_ids:
    ml.quantity = ml.move_id.product_uom_qty
picking.button_validate()
print("Picking validated: %s" % picking.name)

# Create landed cost with 4 cost lines
lc_product = env['product.product'].search([('name', '=', 'Freight')], limit=1) or p1

lc = env['stock.landed.cost'].create({
    'target_model': 'picking',
    'picking_ids': [(6, 0, [picking.id])],
})

cost_items = [
    ('Freight', 'by_quantity', 500.0),
    ('Duty', 'by_quantity', 300.0),
    ('Insurance', 'by_current_cost_price', 200.0),
    ('Clearing Fee', 'equal', 100.0),
]
for name, split, amount in cost_items:
    env['stock.landed.cost.lines'].create({
        'name': name,
        'split_method': split,
        'price_unit': amount,
        'cost_id': lc.id,
        'product_id': lc_product.id,
    })

journal = env['account.journal'].search([
    ('company_id', '=', env.company.id),
    ('type', '=', 'general'),
], limit=1)

lc.compute_landed_cost()
lc.account_journal_id = journal
lc.button_validate()
print("LC validated: %s" % lc.name)

# Export
result = lc.action_export_distribution_excel()
filepath = result['url'].split('filepath=')[1].split('&')[0]
print("Filepath: %s" % filepath)

if os.path.exists(filepath):
    wb = load_workbook(filepath)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print("\n=== Sheet: %s (rows=%d) ===" % (sn, ws.max_row))
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [str(c) if c is not None else '' for c in row]
            print("Row %2d: %s" % (row_idx, ' | '.join(cells)))
