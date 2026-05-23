import os
from openpyxl import load_workbook

# Use the LC with 4 cost lines we created earlier
lc = env['stock.landed.cost'].search([('name', '=', 'LC/2026/0180')], limit=1)
if not lc:
    print("LC not found")
else:
    print("LC: %s | Cost Lines: %d" % (lc.name, len(lc.cost_lines)))
    print("Adj lines: %d" % len(lc.valuation_adjustment_lines))

    result = lc.action_export_distribution_excel()
    filepath = result['url'].split('filepath=')[1].split('&')[0]
    print("Filepath: %s" % filepath)

    wb = load_workbook(filepath)
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn.startswith('LC_'):
            print("\n=== Sheet: %s ===" % sn)
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                cells = [str(c) if c is not None else '' for c in row]
                if any(c for c in cells):
                    print("Row %2d: %s" % (row_idx, ' | '.join(cells)))
