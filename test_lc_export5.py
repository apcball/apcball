import os
from openpyxl import load_workbook

# Find or create LC with multiple cost lines
product_model = env['product.product']
wh = env['stock.warehouse'].search([('company_id', '=', env.company.id)], limit=1)
supplier_loc = env.ref('stock.stock_location_suppliers')

p1 = product_model.search([('default_code', '=', 'TEST-LC-A')], limit=1)
p2 = product_model.search([('default_code', '=', 'TEST-LC-B')], limit=1)

if not p1:
    p1 = product_model.create({
        'name': 'Test LC Report Product A',
        'default_code': 'TEST-LC-A',
        'type': 'product',
        'standard_price': 200,
        'weight': 5.0,
    })
if not p2:
    p2 = product_model.create({
        'name': 'Test LC Report Product B',
        'default_code': 'TEST-LC-B',
        'type': 'product',
        'standard_price': 250,
        'weight': 3.0,
    })

picking = env['stock.picking'].create({
    'picking_type_id': wh.in_type_id.id,
    'location_id': supplier_loc.id,
    'location_dest_id': wh.lot_stock_id.id,
})
for p, qty in [(p1, 10), (p2, 5)]:
    env['stock.move'].create({
        'name': 'Receipt %s' % p.default_code,
        'product_id': p.id,
        'product_uom_qty': qty,
        'product_uom': p.uom_id.id,
        'picking_id': picking.id,
        'location_id': supplier_loc.id,
        'location_dest_id': wh.lot_stock_id.id,
        'price_unit': p.standard_price,
    })
picking.action_confirm()
picking.action_assign()
for ml in picking.move_line_ids:
    ml.quantity = ml.move_id.product_uom_qty
picking.button_validate()
print("Picking: %s" % picking.name)

lc_product = product_model.search([('name', '=', 'Freight')], limit=1) or p1
lc = env['stock.landed.cost'].create({
    'target_model': 'picking',
    'picking_ids': [(6, 0, [picking.id])],
})
for name, split, amount in [
    ('Freight', 'by_quantity', 500.0),
    ('Duty', 'by_quantity', 300.0),
    ('Insurance', 'by_current_cost_price', 200.0),
    ('Clearing Fee', 'equal', 100.0),
]:
    env['stock.landed.cost.lines'].create({
        'name': name,
        'split_method': split,
        'price_unit': amount,
        'cost_id': lc.id,
        'product_id': lc_product.id,
    })

journal = env['account.journal'].search([('company_id', '=', env.company.id), ('type', '=', 'general')], limit=1)
lc.compute_landed_cost()
lc.account_journal_id = journal
lc.button_validate()
print("LC: %s" % lc.name)
print("Adj lines: %d" % len(lc.valuation_adjustment_lines))

# Export
result = lc.action_export_distribution_excel()
filepath = result['url'].split('filepath=')[1].split('&')[0]

wb = load_workbook(filepath)
for sn in wb.sheetnames:
    ws = wb[sn]
    if sn.startswith('LC_'):
        print("\n=== Sheet: %s ===" % sn)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [str(c) if c is not None else '' for c in row]
            if any(c for c in cells):
                print("Row %2d: %s" % (row_idx, ' | '.join(cells)))
