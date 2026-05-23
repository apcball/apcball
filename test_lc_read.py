import os
from openpyxl import load_workbook

filepath = '/tmp/landed_cost_distribution_20260523.xlsx'
if os.path.exists(filepath):
    wb = load_workbook(filepath)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print("\n=== Sheet: %s ===" % sn)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            cells = [str(c) if c is not None else '' for c in row]
            print("Row %2d: %s" % (row_idx, ' | '.join(cells)))
else:
    print("File not found")
